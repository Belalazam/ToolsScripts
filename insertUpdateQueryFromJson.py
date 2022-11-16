import pandas as pd
import openpyxl as opxl;

# enter the name of file location 
dataframe = opxl.load_workbook("C:\\Users\\BelalAzam\\Downloads\\bs_aa_tracker.xlsx")




# things to note before running this script
# this script will take table name from sheet name
# this will generate the query all at once in different-different text file, text file name will be same as sheet name



print("started for ",dataframe.sheetnames)

for i in dataframe.sheetnames:
    f = open(i+".txt","x")
    singleSheet = dataframe[i]
    temp = "insert into "+ i + " ("
    for row in range(0, 1):
        
        for col in singleSheet.iter_cols(0, singleSheet.max_column):
            temp+= str(col[row].value)+ ","
        temp = temp[:-1]  
        temp += ") values ("          
    for row in range(1, singleSheet.max_row):
        buildString = temp
        for col in singleSheet.iter_cols(0, singleSheet.max_column):
            bricks= "'" + str(col[row].value) + "'"
            brickSize = len(bricks)
            gamma = ""
            for j in range(0,brickSize):
                if(j > 0 and j < brickSize-1 and bricks[j-1].isalpha and bricks[j+1].isalpha and bricks[j]=='\''):
                    gamma+='`'
                else :
                    gamma+=bricks[j]
            bricks = gamma
            if(bricks == "'#N/A'" or bricks == "'None'" or bricks == "' '"):
                bricks = "null"
            buildString += bricks + ","
        buildString = buildString[:-1]
        buildString +=  ")"  + ";"                      
        f.write('%s\n \n' % buildString)
        buildString = ""
print("done")
    
