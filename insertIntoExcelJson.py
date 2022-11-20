import pandas as pd
import json 
import openpyxl as opxl;

# this code will pickup json from one excel sheet and refractor all those json variables to column in other excel sheet 


dataframeJson = opxl.load_workbook("C:\\Users\\BelalAzam\\Downloads\\fileExcel.xlsx")  #sheetlocation of json
dataframeSaver = opxl.load_workbook("C:\\Users\\BelalAzam\\Downloads\\fileExcel.xlsx") #sheetlocation of file to save
sheetFromJson = dataframeJson["Sheet1"]
sheetFromSaver = dataframeSaver["Sheet2"]

print("insertion starts")

dict = {}
def dictGet(a):
    dict[0] = " "
    try:
        return dict[a]
    except:
        return dict[0]


a = []
for row in range(0, 1):
    for col in sheetFromSaver.iter_cols(0, sheetFromSaver.max_column):
            a.append(str(col[row].value))

n = sheetFromSaver.max_row
for row in range(0,sheetFromJson.max_row):
    for col in sheetFromJson.iter_cols(0,1):
        s = str(col[row].value)
        dict = json.loads(s)
        
        dict = json.loads(s)
    m = 1
    for i in a :
        sheetFromSaver.cell(row=n+1,column=m,value = str(dictGet(i)))
        m = m+1      
    n = n+1
    dict.clear()


dataframeSaver.save("C:\\Users\\BelalAzam\\Downloads\\fileExcel.xlsx")
print("insertion Ends")
